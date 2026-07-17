import os
import tempfile
from decimal import Decimal
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Asset paths — images embedded in PDF/Excel exports
_ASSETS = os.path.join(os.path.dirname(__file__), 'assets')
LOGO_FULL  = os.path.join(_ASSETS, 'logo_full.png')   # Group Nish logo with tagline
LOGO_MARK  = os.path.join(_ASSETS, 'logo_mark.png')   # Circular mark only (Excel)
STAMP_PATH = os.path.join(_ASSETS, 'stamp.png')       # Company round stamp


def amount_in_words(amount):
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight",
            "Nine", "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
            "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def _below_thousand(n):
        if n == 0: return ""
        elif n < 20: return ones[n] + " "
        elif n < 100: return tens[n // 10] + (" " + ones[n % 10] if n % 10 else "") + " "
        else: return ones[n // 100] + " Hundred " + _below_thousand(n % 100)

    amount = int(round(amount))
    if amount == 0: return "Zero Rupees Only"
    crore = amount // 10000000; amount %= 10000000
    lakh  = amount // 100000;   amount %= 100000
    thou  = amount // 1000;     amount %= 1000
    hund  = amount

    result = ""
    if crore: result += _below_thousand(crore) + "Crore "
    if lakh:  result += _below_thousand(lakh)  + "Lakh "
    if thou:  result += _below_thousand(thou)  + "Thousand "
    if hund:  result += _below_thousand(hund)
    return result.strip() + " Rupees Only"


def generate_ra_excel(ra, project, line_items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RA Bill"

    hdr_font   = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)
    green_fill = PatternFill("solid", fgColor="E1F5EE")
    gray_fill  = PatternFill("solid", fgColor="F1EFE8")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right", vertical="center")
    left   = Alignment(horizontal="left",  vertical="center", wrap_text=True)

    def cell(row, col, value="", bold=False, fill=None, align=center, fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if bold: c.font = Font(bold=True, size=10)
        else:    c.font = Font(size=10)
        if fill: c.fill = fill
        c.alignment = align
        c.border = border
        if fmt:  c.number_format = fmt
        return c

    col_widths = [6, 60, 8, 10, 12, 10, 12, 10, 12, 10, 12, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Logo in Excel (top-left, rows 1-2)
    try:
        from openpyxl.drawing.image import Image as XLImage
        xl_logo = XLImage(LOGO_MARK)
        xl_logo.width  = 65
        xl_logo.height = 65
        ws.add_image(xl_logo, 'A1')
        ws.row_dimensions[1].height = 50
        ws.row_dimensions[2].height = 18
    except Exception:
        pass

    r = 3
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"RA BILL No. {ra.ra_number} — {project.name}")
    c.font = title_font; c.alignment = center; c.fill = green_fill; c.border = border
    r += 1

    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"Invoice No: {ra.invoice_no}   |   Date: {ra.invoice_date}   |   WO: {project.wo_number}")
    c.font = Font(size=10); c.alignment = center; c.border = border
    r += 1

    ws.merge_cells(f"A{r}:F{r+2}")
    c = ws.cell(r, 1, f"SELLER: {project.seller_name}\n{project.seller_address}\nGSTIN: {project.seller_gstin}")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=9); c.border = border

    ws.merge_cells(f"G{r}:M{r+2}")
    c = ws.cell(r, 7, f"BUYER: {project.client_name}\n{project.client_address}\nGSTIN: {project.client_gstin}")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=9); c.border = border
    r += 3

    headers = ["Sr.", "Description", "Unit", "PO Qty", "Rate",
               "Prev Qty", "Prev Amt", "This Qty", "This Amt",
               "Upto Qty", "Upto Amt", "Bal Qty", "Bal Amt"]
    for i, h in enumerate(headers, 1):
        cell(r, i, h, bold=True, fill=gray_fill)
    r += 1

    # Group by Item No. (sr_no). Split items (same sr_no, multiple stages)
    # show ONE description header row, then a sub-row per stage that exists.
    STAGE_LABELS = {"supply": "Supply", "erection": "Installation", "commissioning": "Commissioning"}

    groups = {}
    group_order = []
    for li in line_items:
        key = li["sr_no"]
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(li)

    for sr_no in group_order:
        items = groups[sr_no]
        desc = items[0]["description"][:200]

        if len(items) > 1:
            # Header row — item no + description, spans across the qty/rate columns
            ws.merge_cells(f"B{r}:M{r}")
            cell(r, 1, sr_no, bold=True, fill=gray_fill, align=center)
            cell(r, 2, desc, bold=True, fill=gray_fill, align=left)
            r += 1
            for li in items:
                stage_label = STAGE_LABELS.get(li.get("item_type", ""), li.get("item_type", ""))
                cell(r, 1, "", align=center)
                cell(r, 2, f"    • {stage_label}", align=left)
                cell(r, 3, li["unit"], align=center)
                cell(r, 4, li["po_qty"], align=right, fmt="#,##0.000")
                cell(r, 5, li["rate"], align=right, fmt="#,##0.00")
                cell(r, 6, li["qty_prev"], align=right, fmt="#,##0.000")
                cell(r, 7, li["amount_prev"], align=right, fmt="#,##0.00")
                cell(r, 8, li["qty_this"], align=right, fmt="#,##0.000")
                cell(r, 9, li["amount_this"], align=right, fmt="#,##0.00")
                cell(r, 10, li["qty_upto"], align=right, fmt="#,##0.000")
                cell(r, 11, li["amount_upto"], align=right, fmt="#,##0.00")
                cell(r, 12, li["qty_balance"], align=right, fmt="#,##0.000")
                cell(r, 13, li["amount_balance"], align=right, fmt="#,##0.00")
                r += 1
        else:
            li = items[0]
            cell(r, 1, li["sr_no"], align=center)
            cell(r, 2, li["description"][:200], align=left)
            cell(r, 3, li["unit"], align=center)
            cell(r, 4, li["po_qty"], align=right, fmt="#,##0.000")
            cell(r, 5, li["rate"], align=right, fmt="#,##0.00")
            cell(r, 6, li["qty_prev"], align=right, fmt="#,##0.000")
            cell(r, 7, li["amount_prev"], align=right, fmt="#,##0.00")
            cell(r, 8, li["qty_this"], align=right, fmt="#,##0.000")
            cell(r, 9, li["amount_this"], align=right, fmt="#,##0.00")
            cell(r, 10, li["qty_upto"], align=right, fmt="#,##0.000")
            cell(r, 11, li["amount_upto"], align=right, fmt="#,##0.00")
            cell(r, 12, li["qty_balance"], align=right, fmt="#,##0.000")
            cell(r, 13, li["amount_balance"], align=right, fmt="#,##0.00")
            r += 1

    r += 1
    def summary_row(label, value, bold=False, fill=None):
        nonlocal r
        ws.merge_cells(f"A{r}:J{r}")
        c = ws.cell(r, 1, label)
        c.font = Font(bold=bold, size=10)
        c.alignment = right; c.border = border
        if fill: c.fill = fill
        c2 = ws.cell(r, 11, value)
        ws.merge_cells(f"K{r}:M{r}")
        c2.font = Font(bold=bold, size=10)
        c2.alignment = right; c2.border = border; c2.number_format = "#,##0.00"
        if fill: c2.fill = fill
        r += 1

    summary_row("Supply value (this bill)", float(ra.supply_value_this))
    summary_row("Installation value (this bill)", float(ra.installation_value_this or 0))
    summary_row("Commissioning value (this bill)", float(ra.commissioning_value_this or 0))
    summary_row("Taxable value", float(ra.taxable_value), bold=True)
    if float(ra.igst_amount) > 0:
        summary_row(f"IGST @ {project.igst_rate}%", float(ra.igst_amount))
    if float(ra.cgst_amount) > 0:
        summary_row(f"CGST @ {project.cgst_rate}%", float(ra.cgst_amount))
        summary_row(f"SGST @ {project.sgst_rate}%", float(ra.sgst_amount))
    summary_row("Gross total", float(ra.gross_total), bold=True)
    summary_row(f"Less: Advance recovery ({project.pt_advance_pct}%)", -float(ra.advance_recovery))
    if float(ra.retention_deduction) > 0:
        summary_row(f"Less: Retention ({project.pt_retention_pct}%)", -float(ra.retention_deduction))
    summary_row("Net payable", float(ra.net_payable), bold=True, fill=green_fill)

    r += 1
    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"Amount in words: {amount_in_words(float(ra.net_payable))}")
    c.font = Font(bold=True, size=10); c.alignment = left; c.border = border
    r += 1

    ws.merge_cells(f"A{r}:M{r}")
    c = ws.cell(r, 1, f"HSN/SAC: {project.hsn_sac_code}   |   Place of Supply: {project.place_of_supply}")
    c.font = Font(size=9); c.alignment = center; c.border = border
    r += 2

    # Stamp in signatory area
    try:
        from openpyxl.drawing.image import Image as XLImage
        stamp_xl = XLImage(STAMP_PATH)
        stamp_xl.width  = 75
        stamp_xl.height = 75
        ws.add_image(stamp_xl, f'K{r}')
    except Exception:
        pass

    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(r, 1, "Certified that the particulars given above are true and correct.")
    c.font = Font(size=9); c.alignment = left; c.border = border
    ws.merge_cells(f"H{r}:M{r}")
    c = ws.cell(r, 8, "For NISH TECHNO PROJECTS PRIVATE LIMITED")
    c.font = Font(bold=True, size=10); c.alignment = center; c.border = border
    r += 4
    ws.merge_cells(f"H{r}:M{r}")
    c = ws.cell(r, 8, "Authorised Signatory")
    c.font = Font(size=10); c.alignment = center; c.border = border

    ws.freeze_panes = "A8"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


def generate_ra_pdf(ra, project, line_items):
    """Generate RA Bill PDF using ReportLab — pure Python, no system dependencies."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)

    GREEN      = colors.HexColor("#0F6E56")
    GREEN_FILL = colors.HexColor("#E1F5EE")
    GRAY_FILL  = colors.HexColor("#F1EFE8")
    BORDER     = colors.HexColor("#CCCCCC")

    doc = SimpleDocTemplate(
        tmp.name, pagesize=landscape(A4),
        topMargin=10*mm, bottomMargin=10*mm,
        leftMargin=10*mm, rightMargin=10*mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Heading2"],
                                  alignment=TA_CENTER, textColor=GREEN, fontSize=14, spaceAfter=2)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"],
                                     alignment=TA_CENTER, fontSize=8.5,
                                     textColor=colors.HexColor("#555555"))
    party_style = ParagraphStyle("party", parent=styles["Normal"], fontSize=7.5, leading=10)
    cell_style  = ParagraphStyle("cell", parent=styles["Normal"], fontSize=7, leading=8.5)
    cell_right  = ParagraphStyle("cellr", parent=cell_style, alignment=TA_RIGHT)
    cell_center = ParagraphStyle("cellc", parent=cell_style, alignment=TA_CENTER)
    words_style = ParagraphStyle("words", parent=styles["Normal"], fontSize=9,
                                  fontName="Helvetica-Bold")

    elements = []

    # Logo header
    try:
        logo_img = RLImage(LOGO_FULL, width=55*mm, height=30*mm)
        logo_img.hAlign = 'CENTER'
        elements.append(logo_img)
        elements.append(Spacer(1, 4))
    except Exception:
        pass

    elements.append(Paragraph(f"RA Bill No. {ra.ra_number} — {project.name}", title_style))
    elements.append(Paragraph(
        f"Invoice No: {ra.invoice_no} | Date: {ra.invoice_date} | WO: {project.wo_number} | HSN: {project.hsn_sac_code}",
        subtitle_style))
    elements.append(Spacer(1, 6))

    # Seller / Buyer / Site blocks
    party_data = [[
        Paragraph(f"<b>Seller:</b><br/>{project.seller_name}<br/>{project.seller_address or ''}<br/>GSTIN: {project.seller_gstin or ''}", party_style),
        Paragraph(f"<b>Buyer:</b><br/>{project.client_name}<br/>{project.client_address or ''}<br/>GSTIN: {project.client_gstin or ''}", party_style),
        Paragraph(f"<b>Site:</b><br/>{project.site_name or project.client_name}<br/>{project.site_address or ''}", party_style),
    ]]
    party_table = Table(party_data, colWidths=[90*mm, 90*mm, 90*mm])
    party_table.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID", (0,0), (-1,-1), 0.5, BORDER),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elements.append(party_table)
    elements.append(Spacer(1, 8))

    # Line items table — grouped by Item No. (sr_no).
    # Split BOQ items (same sr_no, multiple stages: supply/erection/commissioning)
    # show ONE description header row, then a sub-row per stage that actually
    # exists (0% stages are never created, so they're naturally absent).
    STAGE_LABELS = {"supply": "Supply", "erection": "Installation", "commissioning": "Commissioning"}

    headers = ["Sr.", "Description", "Unit", "PO Qty", "Rate",
               "Prev Qty", "Prev Amt", "This Qty", "This Amt",
               "Upto Qty", "Upto Amt", "Bal Qty", "Bal Amt"]
    table_data = [[Paragraph(f"<b>{h}</b>", cell_center) for h in headers]]
    header_row_indices = []  # rows to span + shade as group headers
    row_idx = 1  # header row is row 0

    groups = {}
    group_order = []
    for li in line_items:
        key = li["sr_no"]
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(li)

    for sr_no in group_order:
        items = groups[sr_no]
        desc = items[0]["description"][:150]

        if len(items) > 1:
            # Multi-stage grouped item — description header spans the row, then stage sub-rows
            table_data.append([
                Paragraph(f"<b>{sr_no}</b>", cell_style),
                Paragraph(f"<b>{desc}</b>", cell_style),
                "", "", "", "", "", "", "", "", "", "", "",
            ])
            header_row_indices.append(row_idx)
            row_idx += 1
            for li in items:
                stage_label = STAGE_LABELS.get(li.get("item_type", ""), li.get("item_type", ""))
                table_data.append([
                    Paragraph("", cell_center),
                    Paragraph(f"&nbsp;&nbsp;&nbsp;• {stage_label}", cell_style),
                    Paragraph(li["unit"], cell_center),
                    Paragraph(f"{li['po_qty']:,.3f}", cell_right),
                    Paragraph(f"{li['rate']:,.2f}", cell_right),
                    Paragraph(f"{li['qty_prev']:,.3f}", cell_right),
                    Paragraph(f"{li['amount_prev']:,.2f}", cell_right),
                    Paragraph(f"{li['qty_this']:,.3f}", cell_right),
                    Paragraph(f"{li['amount_this']:,.2f}", cell_right),
                    Paragraph(f"{li['qty_upto']:,.3f}", cell_right),
                    Paragraph(f"{li['amount_upto']:,.2f}", cell_right),
                    Paragraph(f"{li['qty_balance']:,.3f}", cell_right),
                    Paragraph(f"{li['amount_balance']:,.2f}", cell_right),
                ])
                row_idx += 1
        else:
            # Single-stage item (old-style manual entry) — unchanged flat row
            li = items[0]
            table_data.append([
                Paragraph(str(li["sr_no"]), cell_center),
                Paragraph(desc, cell_style),
                Paragraph(li["unit"], cell_center),
                Paragraph(f"{li['po_qty']:,.3f}", cell_right),
                Paragraph(f"{li['rate']:,.2f}", cell_right),
                Paragraph(f"{li['qty_prev']:,.3f}", cell_right),
                Paragraph(f"{li['amount_prev']:,.2f}", cell_right),
                Paragraph(f"{li['qty_this']:,.3f}", cell_right),
                Paragraph(f"{li['amount_this']:,.2f}", cell_right),
                Paragraph(f"{li['qty_upto']:,.3f}", cell_right),
                Paragraph(f"{li['amount_upto']:,.2f}", cell_right),
                Paragraph(f"{li['qty_balance']:,.3f}", cell_right),
                Paragraph(f"{li['amount_balance']:,.2f}", cell_right),
            ])
            row_idx += 1

    col_widths = [10*mm, 55*mm, 12*mm, 16*mm, 16*mm, 16*mm, 18*mm, 16*mm, 18*mm, 16*mm, 18*mm, 16*mm, 18*mm]
    table_style_cmds = [
        ("BACKGROUND", (0,0), (-1,0), GREEN_FILL),
        ("GRID", (0,0), (-1,-1), 0.5, BORDER),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
    ]
    for hr in header_row_indices:
        table_style_cmds.append(("SPAN", (1, hr), (12, hr)))
        table_style_cmds.append(("BACKGROUND", (0, hr), (-1, hr), GRAY_FILL))

    items_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle(table_style_cmds))
    elements.append(items_table)
    elements.append(Spacer(1, 10))

    # Summary table
    summary_rows = [
        ["Supply value (this bill)", f"Rs. {float(ra.supply_value_this):,.2f}", False],
        ["Installation value (this bill)", f"Rs. {float(ra.installation_value_this or 0):,.2f}", False],
        ["Commissioning value (this bill)", f"Rs. {float(ra.commissioning_value_this or 0):,.2f}", False],
        ["Taxable value", f"Rs. {float(ra.taxable_value):,.2f}", True],
    ]
    if float(ra.igst_amount) > 0:
        summary_rows.append([f"IGST @ {project.igst_rate}%", f"Rs. {float(ra.igst_amount):,.2f}", False])
    if float(ra.cgst_amount) > 0:
        summary_rows.append([f"CGST @ {project.cgst_rate}%", f"Rs. {float(ra.cgst_amount):,.2f}", False])
        summary_rows.append([f"SGST @ {project.sgst_rate}%", f"Rs. {float(ra.sgst_amount):,.2f}", False])
    summary_rows.append(["Gross total", f"Rs. {float(ra.gross_total):,.2f}", True])
    summary_rows.append([f"Less: Advance recovery ({project.pt_advance_pct}%)", f"Rs. {float(ra.advance_recovery):,.2f}", False])
    if float(ra.retention_deduction) > 0:
        summary_rows.append([f"Less: Retention ({project.pt_retention_pct}%)", f"Rs. {float(ra.retention_deduction):,.2f}", False])
    summary_rows.append(["Net payable", f"Rs. {float(ra.net_payable):,.2f}", True])

    summary_style   = ParagraphStyle("summary", parent=styles["Normal"], fontSize=9)
    summary_style_b = ParagraphStyle("summaryb", parent=summary_style, fontName="Helvetica-Bold")
    summary_right   = ParagraphStyle("summaryr", parent=summary_style, alignment=TA_RIGHT)
    summary_right_b = ParagraphStyle("summaryrb", parent=summary_style_b, alignment=TA_RIGHT)

    summary_data = []
    for label, value, bold in summary_rows:
        summary_data.append([
            Paragraph(label, summary_style_b if bold else summary_style),
            Paragraph(value, summary_right_b if bold else summary_right),
        ])

    summary_table = Table(summary_data, colWidths=[130*mm, 50*mm], hAlign="RIGHT")
    sstyle_cmds = [
        ("GRID", (0,0), (-1,-1), 0.4, BORDER),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0, len(summary_data)-1), (-1, len(summary_data)-1), GREEN_FILL),
    ]
    summary_table.setStyle(TableStyle(sstyle_cmds))
    elements.append(summary_table)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph(
        f"Amount in words: {amount_in_words(float(ra.net_payable))}", words_style))
    elements.append(Spacer(1, 20))

    # Signatory footer with stamp
    try:
        stamp_img = RLImage(STAMP_PATH, width=30*mm, height=30*mm)
    except Exception:
        stamp_img = Paragraph("", party_style)

    footer_data = [[
        Paragraph(f"For {project.client_name}<br/><br/><br/>Authorised Signatory", party_style),
        [Paragraph(f"For {project.seller_name}", party_style),
         stamp_img,
         Paragraph("Authorised Signatory", party_style)],
    ]]
    footer_table = Table(footer_data, colWidths=[140*mm, 140*mm])
    footer_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    elements.append(footer_table)

    doc.build(elements)
    tmp.close()
    return tmp.name


# ─────────────────────────────────────────────────────────────────────────────
# TAX INVOICE GENERATORS
# ─────────────────────────────────────────────────────────────────────────────

def generate_tax_invoice_pdf(ra, project):
    """Generate GST Tax Invoice PDF using ReportLab — pure Python, no system deps."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable,
                                     Image as RLImage)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)

    GREEN      = colors.HexColor("#0F6E56")
    GREEN_FILL = colors.HexColor("#E1F5EE")
    GRAY_FILL  = colors.HexColor("#F1EFE8")
    BORDER     = colors.HexColor("#CCCCCC")

    doc = SimpleDocTemplate(
        tmp.name, pagesize=A4,
        topMargin=10*mm, bottomMargin=10*mm,
        leftMargin=12*mm, rightMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s  = ps("ti", fontSize=13, fontName="Helvetica-Bold", alignment=TA_CENTER, textColor=GREEN)
    hdr_s    = ps("hd", fontSize=8, fontName="Helvetica-Bold")
    cell_s   = ps("ce", fontSize=8, leading=10)
    cell_b   = ps("cb", fontSize=8, fontName="Helvetica-Bold", leading=10)
    cell_r   = ps("cr", fontSize=8, alignment=TA_RIGHT)
    cell_rb  = ps("crb", fontSize=8, fontName="Helvetica-Bold", alignment=TA_RIGHT)
    cell_c   = ps("cc", fontSize=8, alignment=TA_CENTER)
    footer_s = ps("fo", fontSize=8, textColor=colors.HexColor("#444444"))

    elements = []

    # Logo header
    try:
        logo_img = RLImage(LOGO_FULL, width=50*mm, height=27*mm)
        logo_img.hAlign = 'CENTER'
        elements.append(logo_img)
        elements.append(Spacer(1, 4))
    except Exception:
        pass

    # Title — no IRN line
    elements.append(Paragraph("TAX INVOICE — RUNNING ACCOUNT (RA) BILL", title_s))
    elements.append(Spacer(1, 4))

    # Bill meta row
    meta_data = [[
        Paragraph(f"<b>RA Bill No.:</b> {ra.ra_number}", cell_s),
        Paragraph(f"<b>Invoice No.:</b> {ra.invoice_no}", cell_s),
        Paragraph(f"<b>Invoice Date:</b> {ra.invoice_date}", cell_s),
    ],[
        Paragraph(f"<b>W.O. No.:</b> {project.wo_number}", cell_s),
        Paragraph(f"<b>W.O. Date:</b> {project.wo_date or ''}", cell_s),
        Paragraph(f"<b>HSN/SAC:</b> {project.hsn_sac_code}  |  <b>Reverse Charge:</b> No", cell_s),
    ],[
        Paragraph(f"<b>Place of Supply:</b> {project.place_of_supply}", cell_s),
        Paragraph("", cell_s),
        Paragraph("", cell_s),
    ]]
    meta_tbl = Table(meta_data, colWidths=[62*mm, 62*mm, 62*mm])
    meta_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 5))

    # Party details
    def party_block(title, name, addr, gstin, pan=None):
        lines = [f"<b>{title}</b>", f"<b>{name}</b>", addr or ""]
        if gstin: lines.append(f"GST No: {gstin}")
        if pan:   lines.append(f"PAN No: {pan}")
        return Paragraph("<br/>".join(lines), cell_s)

    party_data = [[
        party_block("SELLER", project.seller_name, project.seller_address,
                    project.seller_gstin, getattr(project, "seller_pan", None)),
        party_block("BUYER", project.client_name, project.client_address,
                    project.client_gstin, getattr(project, "client_pan", None)),
        party_block("CONSIGNEE / SITE", project.site_name or project.client_name,
                    project.site_address or "", None),
    ]]
    party_tbl = Table(party_data, colWidths=[62*mm, 62*mm, 62*mm])
    party_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(party_tbl)
    elements.append(Spacer(1, 6))

    # Abstract of Bill
    elements.append(Paragraph("<b>ABSTRACT OF BILL (Taxable Values)</b>", hdr_s))
    elements.append(Spacer(1, 3))

    order_val_p1 = float(getattr(project, "wo_value_supply", 0) or 0)
    order_val_p2 = float(getattr(project, "wo_value_ec", 0) or 0)
    order_total  = order_val_p1 + order_val_p2
    sup_prev = float(ra.supply_value_prev); sup_this = float(ra.supply_value_this)
    sup_upto = float(ra.supply_value_upto)
    ec_prev  = float(ra.ec_value_prev);  ec_this  = float(ra.ec_value_this)
    ec_upto  = float(ra.ec_value_upto)
    tax_val  = float(ra.taxable_value); igst = float(ra.igst_amount)
    cgst = float(ra.cgst_amount); sgst = float(ra.sgst_amount)
    gross    = float(ra.gross_total); adv_rec = float(ra.advance_recovery)
    ret_ded  = float(ra.retention_deduction) if hasattr(ra, "retention_deduction") else 0
    net_pay  = float(ra.net_payable)

    def inr(v): return f"Rs. {v:,.2f}"

    abs_headers = [
        Paragraph("<b>Part</b>", cell_b),
        Paragraph("<b>Order Value</b>", cell_rb),
        Paragraph("<b>Up-to-Date Billed</b>", cell_rb),
        Paragraph("<b>Previous Billed</b>", cell_rb),
        Paragraph("<b>THIS BILL</b>", cell_rb),
    ]
    abs_data = [
        abs_headers,
        [Paragraph("Part 1 — Balance Supply (SITC)", cell_s),
         Paragraph(inr(order_val_p1), cell_r),
         Paragraph(inr(sup_upto), cell_r),
         Paragraph(inr(sup_prev), cell_r),
         Paragraph(inr(sup_this), cell_r)],
        [Paragraph("Part 2 — Installation & Commissioning", cell_s),
         Paragraph(inr(order_val_p2), cell_r),
         Paragraph(inr(ec_upto), cell_r),
         Paragraph(inr(ec_prev), cell_r),
         Paragraph(inr(ec_this), cell_r)],
        [Paragraph("<b>TOTAL TAXABLE AMOUNT</b>", cell_b),
         Paragraph(inr(order_total), cell_rb),
         Paragraph(inr(sup_upto + ec_upto), cell_rb),
         Paragraph(inr(sup_prev + ec_prev), cell_rb),
         Paragraph(inr(tax_val), cell_rb)],
    ]
    abs_tbl = Table(abs_data, colWidths=[65*mm, 30*mm, 35*mm, 30*mm, 26*mm])
    abs_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), GRAY_FILL),
        ("BACKGROUND",  (0,3), (-1,3), GREEN_FILL),
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING",(0,0), (-1,-1), 4),
    ]))
    elements.append(abs_tbl)
    elements.append(Spacer(1, 4))

    # Tax & Summary block
    def summary_row(label, value, bold=False):
        ls = cell_b if bold else cell_s
        rs = cell_rb if bold else cell_r
        return [Paragraph(f"<b>{label}</b>" if bold else label, ls),
                Paragraph(f"<b>{inr(value)}</b>" if bold else inr(value), rs)]

    sum_data = []
    if igst > 0: sum_data.append(summary_row(f"IGST @ {project.igst_rate}%", igst))
    if cgst > 0:
        sum_data.append(summary_row(f"CGST @ {project.cgst_rate}%", cgst))
        sum_data.append(summary_row(f"SGST @ {project.sgst_rate}%", sgst))
    sum_data.append(summary_row("TOTAL INVOICE VALUE (incl. GST)", gross, bold=True))
    sum_data.append(summary_row(
        f"Less: Advance adjusted ({project.pt_advance_pct}% of Part-1 this bill)", adv_rec))
    if ret_ded > 0:
        sum_data.append(summary_row(f"Less: Retention ({project.pt_retention_pct}%)", ret_ded))
    sum_data.append(summary_row("NET AMOUNT RECEIVABLE", net_pay, bold=True))

    sum_tbl = Table(sum_data, colWidths=[130*mm, 56*mm], hAlign="RIGHT")
    sstyle = [
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING",(0,0), (-1,-1), 5),
        ("BACKGROUND",  (0, len(sum_data)-1), (-1, len(sum_data)-1), GREEN_FILL),
    ]
    sum_tbl.setStyle(TableStyle(sstyle))
    elements.append(sum_tbl)
    elements.append(Spacer(1, 6))

    # Amount in words
    elements.append(Paragraph(
        f"<b>Invoice Value in Words:</b> {amount_in_words(gross)} (incl. GST)", cell_b))
    elements.append(Paragraph(
        f"<b>Net Amount Receivable in Words:</b> {amount_in_words(net_pay)}", cell_b))
    elements.append(Spacer(1, 8))

    # Payment terms + signatory with stamp
    pay_terms = [
        getattr(project, "pt_part1_terms",
                "20% advance (received) + 80% through 60-days usance LC from shipment date"),
        getattr(project, "pt_part2_terms",
                "80% on installation + 20% on commissioning, pro-rata, 15 days from RA bill certification"),
    ]

    try:
        ti_stamp = RLImage(STAMP_PATH, width=28*mm, height=28*mm)
    except Exception:
        ti_stamp = Paragraph("", footer_s)

    footer_data = [[
        Paragraph("<br/>".join([
            "<b>Payment Terms:</b>",
            f"Part-1: {pay_terms[0]}",
            f"Part-2: {pay_terms[1]}",
            "",
            "Certified that the particulars given above are true and correct and the amount",
            "claimed is as per actual work executed.  E. & O. E.",
        ]), footer_s),
        [Paragraph("For <b>NISH TECHNO PROJECTS PRIVATE LIMITED</b>", footer_s),
         ti_stamp,
         Paragraph("Authorised Signatory", footer_s)],
    ]]
    footer_table = Table(footer_data, colWidths=[100*mm, 86*mm])
    footer_table.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
    ]))
    elements.append(footer_table)

    doc.build(elements)
    tmp.close()
    return tmp.name


def generate_tax_invoice_excel(ra, project):
    """Generate GST Tax Invoice as formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"

    thin  = Side(style="thin")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L     = Alignment(horizontal="left", vertical="center", wrap_text=True)
    R     = Alignment(horizontal="right", vertical="center")
    GREEN_FILL = PatternFill("solid", fgColor="E1F5EE")
    GRAY_FILL  = PatternFill("solid", fgColor="F1EFE8")
    TEAL_FILL  = PatternFill("solid", fgColor="0F6E56")

    col_widths = [4, 28, 18, 18, 18, 18, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def sc(r, c, val="", bold=False, fill=None, align=L, fmt=None, size=10, color="000000"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, size=size, name="Arial", color=color)
        if fill: cell.fill = fill
        cell.alignment = align
        cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    def merge(r, c1, c2, val="", bold=False, fill=None, align=C, size=10, color="000000"):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        return sc(r, c1, val, bold, fill, align, size=size, color=color)

    # Logo in Excel (top-left)
    try:
        from openpyxl.drawing.image import Image as XLImage
        xl_logo_ti = XLImage(LOGO_MARK)
        xl_logo_ti.width  = 60
        xl_logo_ti.height = 60
        ws.add_image(xl_logo_ti, 'A1')
        ws.row_dimensions[1].height = 48
        ws.row_dimensions[2].height = 16
    except Exception:
        pass

    r = 3
    merge(r, 1, 6, "TAX INVOICE — RUNNING ACCOUNT (RA) BILL",
          bold=True, fill=TEAL_FILL, size=13, color="FFFFFF")
    ws.row_dimensions[r].height = 24; r += 1
    # IRN line removed
    ws.row_dimensions[r].height = 4; r += 1

    sc(r, 2, f"RA Bill No.: {ra.ra_number}", bold=True)
    sc(r, 3, f"Invoice No.: {ra.invoice_no}", bold=True)
    sc(r, 4, f"Invoice Date: {ra.invoice_date}", bold=True)
    sc(r, 5, f"HSN/SAC: {project.hsn_sac_code}", bold=True)
    r += 1
    sc(r, 2, f"W.O. No.: {project.wo_number}")
    sc(r, 3, f"Place of Supply: {project.place_of_supply}")
    sc(r, 4, "Reverse Charge: No")
    sc(r, 5, f"W.O. Date: {project.wo_date or ''}")
    r += 1

    ws.row_dimensions[r].height = 10; r += 1

    headers = ["SELLER", "BUYER", "CONSIGNEE / SITE"]
    for i, h in enumerate(headers):
        sc(r, 2 + i, h, bold=True, fill=GRAY_FILL, align=C)
    r += 1

    party_lines = [
        [project.seller_name, project.client_name, project.site_name or project.client_name],
        [project.seller_address or "", project.client_address or "", project.site_address or ""],
        [f"GSTIN: {project.seller_gstin or ''}", f"GSTIN: {project.client_gstin or ''}", ""],
    ]
    for line in party_lines:
        for i, val in enumerate(line):
            sc(r, 2 + i, val, align=L, size=9)
        ws.row_dimensions[r].height = 20
        r += 1

    ws.row_dimensions[r].height = 10; r += 1

    merge(r, 1, 6, "ABSTRACT OF BILL (Taxable Values, Rs.)", bold=True, fill=GRAY_FILL, align=L)
    r += 1

    abs_hdrs = ["Part", "Order Value", "Up-to-Date Billed", "Previous Billed", "THIS BILL", ""]
    for i, h in enumerate(abs_hdrs[:-1], 2):
        sc(r, i, h, bold=True, fill=GRAY_FILL, align=C)
    r += 1

    order_val_p1 = float(getattr(project, "wo_value_supply", 0) or 0)
    order_val_p2 = float(getattr(project, "wo_value_ec", 0) or 0)
    sup_prev = float(ra.supply_value_prev); sup_this = float(ra.supply_value_this)
    sup_upto = float(ra.supply_value_upto)
    ec_prev  = float(ra.ec_value_prev);  ec_this  = float(ra.ec_value_this)
    ec_upto  = float(ra.ec_value_upto)
    tax_val  = float(ra.taxable_value); igst = float(ra.igst_amount)
    cgst = float(ra.cgst_amount); sgst = float(ra.sgst_amount)
    gross    = float(ra.gross_total); adv_rec = float(ra.advance_recovery)
    ret_ded  = float(ra.retention_deduction) if hasattr(ra, "retention_deduction") else 0
    net_pay  = float(ra.net_payable)
    FMT = "#,##0.00"

    rows_data = [
        ("Part 1 — Balance Supply (SITC)", order_val_p1, sup_upto, sup_prev, sup_this),
        ("Part 2 — Installation & Commissioning", order_val_p2, ec_upto, ec_prev, ec_this),
        ("TOTAL TAXABLE AMOUNT", order_val_p1+order_val_p2,
         sup_upto+ec_upto, sup_prev+ec_prev, tax_val),
    ]
    for i, (label, *vals) in enumerate(rows_data):
        fill = GREEN_FILL if i == 2 else None
        bold = i == 2
        sc(r, 2, label, bold=bold, fill=fill, align=L)
        for j, v in enumerate(vals):
            sc(r, 3+j, v, bold=bold, fill=fill, align=R, fmt=FMT)
        r += 1

    ws.row_dimensions[r].height = 8; r += 1

    tax_rows = []
    if igst > 0: tax_rows.append((f"IGST @ {project.igst_rate}%", igst, False))
    if cgst > 0:
        tax_rows.append((f"CGST @ {project.cgst_rate}%", cgst, False))
        tax_rows.append((f"SGST @ {project.sgst_rate}%", sgst, False))
    tax_rows.append(("TOTAL INVOICE VALUE (incl. GST)", gross, True))
    tax_rows.append((f"Less: Advance adjusted ({project.pt_advance_pct}% of Part-1 this bill)",
                     adv_rec, False))
    if ret_ded > 0:
        tax_rows.append((f"Less: Retention ({project.pt_retention_pct}%)", ret_ded, False))
    tax_rows.append(("NET AMOUNT RECEIVABLE", net_pay, True))

    for label, val, bold in tax_rows:
        fill = GREEN_FILL if bold and label.startswith("NET") else (GRAY_FILL if bold else None)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        sc(r, 2, label, bold=bold, fill=fill, align=L)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        sc(r, 5, val, bold=bold, fill=fill, align=R, fmt=FMT)
        r += 1

    ws.row_dimensions[r].height = 8; r += 1

    merge(r, 1, 6, f"Invoice Value in Words: {amount_in_words(gross)} (incl. GST)",
          bold=True, align=L, size=9)
    r += 1
    merge(r, 1, 6, f"Net Amount Receivable in Words: {amount_in_words(net_pay)}",
          bold=True, align=L, size=9)
    r += 1; ws.row_dimensions[r].height = 8; r += 1

    sc(r, 2, "Certified that the particulars given above are true and correct "
             "and the amount claimed is as per actual work executed.  E. & O. E.", size=9)
    sc(r, 5, "For NISH TECHNO PROJECTS PRIVATE LIMITED", bold=True)
    r += 1

    # Stamp in signatory area
    try:
        from openpyxl.drawing.image import Image as XLImage
        stamp_xl = XLImage(STAMP_PATH)
        stamp_xl.width  = 70
        stamp_xl.height = 70
        ws.add_image(stamp_xl, f'E{r}')
    except Exception:
        pass
    r += 3
    sc(r, 5, "Authorised Signatory")

    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(out.name)
    out.close()
    return out.name


def generate_reconciliation_excel(recon_items, project):
    """Export Reconciliation data as formatted Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    thin = Side(style="thin")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L    = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    R    = Alignment(horizontal="right", vertical="center")
    TEAL = PatternFill("solid", fgColor="0F6E56")
    GRAY = PatternFill("solid", fgColor="F1EFE8")
    GRN  = PatternFill("solid", fgColor="E1F5EE")

    col_ws = [12, 6, 45, 14, 8, 16, 16, 16, 18, 35]
    for i, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def sc(r, c, val="", bold=False, fill=None, align=L, fmt=None, size=9):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, size=size, name="Arial",
                         color="FFFFFF" if fill == TEAL else "000000")
        if fill: cell.fill = fill
        cell.alignment = align
        cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    r = 1
    ws.merge_cells(f"A{r}:J{r}")
    sc(r, 1,
       f"RECONCILIATION — OLD WO (Gharpure J-2209/WO-249 Amd-2) vs "
       f"NEW PO (BEIL WO-249 Amd-3)",
       bold=True, fill=TEAL, align=C, size=11)
    ws.row_dimensions[r].height = 22; r += 1

    ws.merge_cells(f"A{r}:J{r}")
    sc(r, 1,
       "Frozen audit record. Old-WO billing stays with Gharpure account; "
       "this order covers only balance supply (Part 1) and pending I&C (Part 2) "
       "at re-negotiated rates.",
       fill=GRAY, align=L, size=8)
    ws.row_dimensions[r].height = 18; r += 1

    hdrs = ["Site", "Old Sr.", "Description", "Old Rate", "Old Qty",
            "Billed Supply (Rs.)", "Billed Install (Rs.)", "Billed Comm (Rs.)",
            "Total Billed Old WO (Rs.)", "Disposition in New PO"]
    for i, h in enumerate(hdrs, 1):
        sc(r, i, h, bold=True, fill=GRAY, align=C)
    ws.row_dimensions[r].height = 30; r += 1

    FMT = "#,##0.00"
    for item in recon_items:
        disp = item.get("disposition", "")
        fill = GRN if "Closed" in disp else (GRAY if "DESCOPED" in disp else None)
        sc(r, 1, item.get("site", ""), fill=fill, align=C)
        sc(r, 2, item.get("old_sr", ""), fill=fill, align=C)
        sc(r, 3, item.get("description", ""), fill=fill)
        sc(r, 4, item.get("old_rate", 0), fill=fill, align=R, fmt=FMT)
        sc(r, 5, item.get("old_qty", 0), fill=fill, align=C)
        sc(r, 6, item.get("billed_supply", 0), fill=fill, align=R, fmt=FMT)
        sc(r, 7, item.get("billed_install", 0), fill=fill, align=R, fmt=FMT)
        sc(r, 8, item.get("billed_comm", 0), fill=fill, align=R, fmt=FMT)
        sc(r, 9, item.get("total_billed", 0), fill=fill, align=R, fmt=FMT)
        sc(r, 10, disp, fill=fill)
        ws.row_dimensions[r].height = 18
        r += 1

    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(out.name)
    out.close()
    return out.name


def generate_challan_pdf(dn, project):
    """Generate Delivery Challan PDF for a dispatch note using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)

    GREEN      = colors.HexColor("#0F6E56")
    GREEN_FILL = colors.HexColor("#E1F5EE")
    GRAY_FILL  = colors.HexColor("#F1EFE8")
    BORDER     = colors.HexColor("#CCCCCC")

    doc = SimpleDocTemplate(
        tmp.name, pagesize=A4,
        topMargin=10*mm, bottomMargin=15*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s  = ps("ti", fontSize=14, fontName="Helvetica-Bold",
                   alignment=TA_CENTER, textColor=GREEN, spaceAfter=2)
    sub_s    = ps("su", fontSize=9, alignment=TA_CENTER,
                   textColor=colors.HexColor("#555555"))
    cell_s   = ps("ce", fontSize=9, leading=12)
    cell_b   = ps("cb", fontSize=9, fontName="Helvetica-Bold", leading=12)
    cell_r   = ps("cr", fontSize=9, alignment=TA_RIGHT)
    cell_c   = ps("cc", fontSize=9, alignment=TA_CENTER)
    footer_s = ps("fo", fontSize=8, textColor=colors.HexColor("#444444"))
    note_s   = ps("no", fontSize=7.5, textColor=colors.HexColor("#666666"), leading=10)

    elements = []

    # Logo
    try:
        logo_img = RLImage(LOGO_FULL, width=45*mm, height=24*mm)
        logo_img.hAlign = "CENTER"
        elements.append(logo_img)
        elements.append(Spacer(1, 3))
    except Exception:
        pass

    elements.append(Paragraph("DELIVERY CHALLAN", title_s))
    elements.append(Paragraph("(For Material Outward — Not a Tax Invoice)", sub_s))
    elements.append(Spacer(1, 6))

    # Challan meta
    boq = dn.boq_item
    rate = float(boq.rate) if boq else 0
    qty  = float(dn.qty_dispatched)
    amount = rate * qty

    meta_data = [[
        Paragraph(f"<b>Challan No.:</b> {dn.dn_number}", cell_s),
        Paragraph(f"<b>Date:</b> {dn.dispatch_date}", cell_s),
        Paragraph(f"<b>WO No.:</b> {project.wo_number or ''}", cell_s),
    ],[
        Paragraph(f"<b>Vehicle No.:</b> {dn.vehicle_no or '—'}", cell_s),
        Paragraph(f"<b>Driver:</b> {dn.driver_name or '—'}", cell_s),
        Paragraph(f"<b>LR No.:</b> {dn.lr_number or '—'}", cell_s),
    ]]
    meta_tbl = Table(meta_data, colWidths=[60*mm, 60*mm, 60*mm])
    meta_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 5))

    # From / To party blocks
    from_to = [[
        Paragraph(
            f"<b>FROM (Consignor):</b><br/>"
            f"<b>{project.seller_name}</b><br/>"
            f"{project.seller_address or ''}<br/>"
            f"GSTIN: {project.seller_gstin or ''}",
            cell_s),
        Paragraph(
            f"<b>TO (Consignee):</b><br/>"
            f"<b>{project.client_name}</b><br/>"
            f"{project.site_name or project.client_name}<br/>"
            f"{project.site_address or ''}<br/>"
            f"Site: {dn.site_destination or ''}",
            cell_s),
    ]]
    ft_tbl = Table(from_to, colWidths=[90*mm, 90*mm])
    ft_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(ft_tbl)
    elements.append(Spacer(1, 6))

    # Item table
    elements.append(Paragraph("<b>MATERIAL DETAILS</b>", cell_b))
    elements.append(Spacer(1, 3))

    item_headers = [
        Paragraph("<b>Sr.</b>", cell_c),
        Paragraph("<b>Item No.</b>", cell_c),
        Paragraph("<b>Description</b>", cell_b),
        Paragraph("<b>HSN/SAC</b>", cell_c),
        Paragraph("<b>Qty</b>", cell_c),
        Paragraph("<b>Unit</b>", cell_c),
        Paragraph("<b>Rate (Rs.)</b>", cell_r),
        Paragraph("<b>Amount (Rs.)</b>", cell_r),
    ]
    item_rows = [item_headers, [
        Paragraph("1", cell_c),
        Paragraph(boq.sr_no if boq else "", cell_c),
        Paragraph(boq.description[:200] if boq else "", cell_s),
        Paragraph(project.hsn_sac_code or "9954", cell_c),
        Paragraph(str(qty), cell_c),
        Paragraph(boq.unit if boq else "", cell_c),
        Paragraph(f"{rate:,.2f}", cell_r),
        Paragraph(f"{amount:,.2f}", cell_r),
    ], [
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph("<b>TOTAL</b>", cell_b),
        Paragraph("", cell_c),
        Paragraph(f"<b>{qty}</b>", cell_c),
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph(f"<b>{amount:,.2f}</b>", cell_r),
    ]]

    item_tbl = Table(item_rows, colWidths=[8*mm, 22*mm, 60*mm, 18*mm, 14*mm, 14*mm, 22*mm, 22*mm])
    item_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), GRAY_FILL),
        ("BACKGROUND",  (0,2), (-1,2), GREEN_FILL),
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING",(0,0), (-1,-1), 3),
    ]))
    elements.append(item_tbl)
    elements.append(Spacer(1, 6))

    # Amount in words
    elements.append(Paragraph(
        f"<b>Amount in words:</b> {amount_in_words(amount)} (excl. GST)", cell_b))

    if dn.remarks:
        elements.append(Spacer(1, 3))
        elements.append(Paragraph(f"<b>Remarks:</b> {dn.remarks}", cell_s))

    elements.append(Spacer(1, 8))

    # Note
    elements.append(Paragraph(
        "Note: This is a Delivery Challan only. This is NOT a Tax Invoice. "
        "GST is not applicable on this document. Invoice will be raised separately.",
        note_s))
    elements.append(Spacer(1, 20))

    # Signatory
    try:
        stamp_img = RLImage(STAMP_PATH, width=28*mm, height=28*mm)
    except Exception:
        stamp_img = Paragraph("", footer_s)

    sign_data = [[
        Paragraph(
            "Received the above material in good condition.<br/><br/><br/>"
            "Receiver's Signature & Stamp:<br/><br/><br/>"
            "Name: ________________  Date: ________________",
            footer_s),
        [Paragraph(f"For {project.seller_name}", footer_s),
         stamp_img,
         Paragraph("Authorised Signatory", footer_s)],
    ]]
    sign_tbl = Table(sign_data, colWidths=[95*mm, 85*mm])
    sign_tbl.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
    ]))
    elements.append(sign_tbl)

    doc.build(elements)
    tmp.close()
    return tmp.name


def generate_po_invoice_pdf(inv, project):
    """Generate PO Invoice PDF using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    GREEN      = colors.HexColor("#0F6E56")
    GREEN_FILL = colors.HexColor("#E1F5EE")
    GRAY_FILL  = colors.HexColor("#F1EFE8")
    BORDER     = colors.HexColor("#CCCCCC")

    doc = SimpleDocTemplate(
        tmp.name, pagesize=A4,
        topMargin=10*mm, bottomMargin=15*mm,
        leftMargin=15*mm, rightMargin=15*mm,
    )

    styles = getSampleStyleSheet()
    def ps(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s  = ps("ti", fontSize=13, fontName="Helvetica-Bold",
                   alignment=TA_CENTER, textColor=GREEN)
    sub_s    = ps("su", fontSize=9, alignment=TA_CENTER,
                   textColor=colors.HexColor("#555555"))
    cell_s   = ps("ce", fontSize=8.5, leading=11)
    cell_b   = ps("cb", fontSize=8.5, fontName="Helvetica-Bold", leading=11)
    cell_r   = ps("cr", fontSize=8.5, alignment=TA_RIGHT)
    cell_rb  = ps("crb", fontSize=8.5, fontName="Helvetica-Bold", alignment=TA_RIGHT)
    cell_c   = ps("cc", fontSize=8.5, alignment=TA_CENTER)
    footer_s = ps("fo", fontSize=8, textColor=colors.HexColor("#444444"))

    elements = []

    # Logo
    try:
        logo_img = RLImage(LOGO_FULL, width=50*mm, height=27*mm)
        logo_img.hAlign = "CENTER"
        elements.append(logo_img)
        elements.append(Spacer(1, 4))
    except Exception:
        pass

    elements.append(Paragraph("TAX INVOICE — PURCHASE ORDER", title_s))
    elements.append(Spacer(1, 4))

    # Meta
    meta_data = [[
        Paragraph(f"<b>Invoice No.:</b> {inv.invoice_no}", cell_s),
        Paragraph(f"<b>Invoice Date:</b> {inv.invoice_date}", cell_s),
        Paragraph(f"<b>PO No.:</b> {project.wo_number or ''}", cell_s),
    ],[
        Paragraph(f"<b>HSN/SAC:</b> {project.hsn_sac_code or '9954'}", cell_s),
        Paragraph(f"<b>Place of Supply:</b> {project.place_of_supply or ''}", cell_s),
        Paragraph("<b>Reverse Charge:</b> No", cell_s),
    ]]
    meta_tbl = Table(meta_data, colWidths=[60*mm, 60*mm, 60*mm])
    meta_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(meta_tbl)
    elements.append(Spacer(1, 5))

    # Party blocks
    party_data = [[
        Paragraph(
            f"<b>SELLER:</b><br/><b>{project.seller_name}</b><br/>"
            f"{project.seller_address or ''}<br/>"
            f"GSTIN: {project.seller_gstin or ''}", cell_s),
        Paragraph(
            f"<b>BUYER:</b><br/><b>{project.client_name}</b><br/>"
            f"{project.client_address or ''}<br/>"
            f"GSTIN: {project.client_gstin or ''}", cell_s),
        Paragraph(
            f"<b>DELIVERY SITE:</b><br/>"
            f"{project.site_name or project.client_name}<br/>"
            f"{project.site_address or ''}", cell_s),
    ]]
    party_tbl = Table(party_data, colWidths=[60*mm, 60*mm, 60*mm])
    party_tbl.setStyle(TableStyle([
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",  (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(party_tbl)
    elements.append(Spacer(1, 6))

    # Line items table
    elements.append(Paragraph("<b>ITEM DETAILS</b>", cell_b))
    elements.append(Spacer(1, 3))

    hdrs = ["Sr.", "DC No.", "Item No.", "Description", "HSN",
            "Qty", "Unit", "Rate (Rs.)", "Amount (Rs.)"]
    item_data = [[Paragraph(f"<b>{h}</b>", cell_c) for h in hdrs]]

    subtotal = float(inv.subtotal or 0)
    for idx, li in enumerate(inv.items, 1):
        item_data.append([
            Paragraph(str(idx), cell_c),
            Paragraph(li.dn.dn_number if li.dn else "", cell_s),
            Paragraph(li.boq_item.sr_no if li.boq_item else "", cell_c),
            Paragraph((li.boq_item.description[:100] if li.boq_item else ""), cell_s),
            Paragraph(li.hsn_code or "", cell_c),
            Paragraph(f"{float(li.qty):,.3f}", cell_r),
            Paragraph(li.boq_item.unit if li.boq_item else "", cell_c),
            Paragraph(f"{float(li.rate):,.2f}", cell_r),
            Paragraph(f"{float(li.amount):,.2f}", cell_r),
        ])

    # Total row
    item_data.append([
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph("<b>TOTAL</b>", cell_b),
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph("", cell_c),
        Paragraph(f"<b>{subtotal:,.2f}</b>", cell_rb),
    ])

    col_w = [8*mm, 22*mm, 18*mm, 55*mm, 14*mm, 14*mm, 10*mm, 20*mm, 22*mm]
    items_tbl = Table(item_data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), GRAY_FILL),
        ("BACKGROUND",  (0,-1),(-1,-1), GREEN_FILL),
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING",(0,0), (-1,-1), 3),
    ]))
    elements.append(items_tbl)
    elements.append(Spacer(1, 6))

    # Summary
    igst = float(inv.igst_amount or 0)
    cgst = float(inv.cgst_amount or 0)
    sgst = float(inv.sgst_amount or 0)
    gross = float(inv.gross_total or 0)

    sum_rows = []
    sum_rows.append(["Subtotal (excl. GST)", f"Rs. {subtotal:,.2f}", False])
    if igst > 0:
        sum_rows.append([f"IGST @ {project.igst_rate}%", f"Rs. {igst:,.2f}", False])
    if cgst > 0:
        sum_rows.append([f"CGST @ {project.cgst_rate}%", f"Rs. {cgst:,.2f}", False])
        sum_rows.append([f"SGST @ {project.sgst_rate}%", f"Rs. {sgst:,.2f}", False])
    sum_rows.append(["TOTAL INVOICE VALUE (incl. GST)", f"Rs. {gross:,.2f}", True])

    def ps_s(bold): return cell_rb if bold else cell_r
    def ps_l(bold): return cell_b if bold else cell_s

    sum_data = [[Paragraph(l, ps_l(b)), Paragraph(v, ps_s(b))]
                for l, v, b in sum_rows]
    sum_tbl = Table(sum_data, colWidths=[120*mm, 50*mm], hAlign="RIGHT")
    sum_style = [
        ("BOX",         (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",   (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING",(0,0), (-1,-1), 5),
        ("BACKGROUND",  (0,len(sum_data)-1),(-1,len(sum_data)-1), GREEN_FILL),
    ]
    sum_tbl.setStyle(TableStyle(sum_style))
    elements.append(sum_tbl)
    elements.append(Spacer(1, 6))

    # Amount in words
    elements.append(Paragraph(
        f"<b>Amount in Words:</b> {amount_in_words(gross)} (incl. GST)", cell_b))
    elements.append(Spacer(1, 20))

    # Signatory
    try:
        stamp_img = RLImage(STAMP_PATH, width=28*mm, height=28*mm)
    except Exception:
        stamp_img = Paragraph("", footer_s)

    sign_data = [[
        Paragraph(
            "Certified that the particulars given above are true and correct.<br/><br/>"
            "E. & O. E.", footer_s),
        [Paragraph(f"For {project.seller_name}", footer_s),
         stamp_img,
         Paragraph("Authorised Signatory", footer_s)],
    ]]
    sign_tbl = Table(sign_data, colWidths=[95*mm, 85*mm])
    sign_tbl.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("BOX",          (0,0), (-1,-1), 0.5, BORDER),
        ("INNERGRID",    (0,0), (-1,-1), 0.3, BORDER),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
    ]))
    elements.append(sign_tbl)

    doc.build(elements)
    tmp.close()
    return tmp.name
