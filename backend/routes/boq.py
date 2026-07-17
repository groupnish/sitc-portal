from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from models.models import (Project, BOQItem, User, GRN, DispatchNote,
                            SiteProgress, RABill, RABillLine, Notification, db)
from services.notifications import (notify_grn_created, notify_dispatch_created,
                                     notify_progress_updated, notify_ra_generated)
from services.export import generate_ra_excel, generate_ra_pdf
from datetime import date, datetime
from decimal import Decimal
import json, os

# ── helpers ──────────────────────────────────────────────────────────────────

def admin_required():
    claims = get_jwt()
    if claims.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403
    return None

def current_user():
    return User.query.get(int(get_jwt_identity()))

# ── Projects ──────────────────────────────────────────────────────────────────

projects_bp = Blueprint("projects", __name__)

@projects_bp.route("/", methods=["GET"])
@jwt_required()
def list_projects():
    projects = Project.query.order_by(Project.created_at.desc()).all()
    return jsonify([p.to_dict() for p in projects])

@projects_bp.route("/", methods=["POST"])
@jwt_required()
def create_project():
    err = admin_required()
    if err: return err
    data = request.get_json()
    p = Project(**{k: v for k, v in data.items() if hasattr(Project, k)})
    if data.get("wo_date"): p.wo_date = date.fromisoformat(data["wo_date"])
    db.session.add(p); db.session.commit()
    return jsonify(p.to_dict()), 201

@projects_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def get_project(pid):
    p = Project.query.get_or_404(pid)
    return jsonify(p.to_dict())

@projects_bp.route("/<int:pid>", methods=["PUT"])
@jwt_required()
def update_project(pid):
    err = admin_required()
    if err: return err
    p = Project.query.get_or_404(pid)
    data = request.get_json()
    for k, v in data.items():
        if hasattr(p, k) and k not in ["id","created_at"]:
            if k == "wo_date" and v: setattr(p, k, date.fromisoformat(v))
            else: setattr(p, k, v)
    db.session.commit()
    return jsonify(p.to_dict())

# ── BOQ ──────────────────────────────────────────────────────────────────────

boq_bp = Blueprint("boq", __name__)

@boq_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def list_boq(pid):
    items = BOQItem.query.filter_by(project_id=pid, is_active=True)\
                         .order_by(BOQItem.sort_order, BOQItem.sr_no).all()
    return jsonify([i.to_dict() for i in items])

@boq_bp.route("/<int:pid>", methods=["POST"])
@jwt_required()
def add_boq_item(pid):
    err = admin_required()
    if err: return err
    data = request.get_json()
    item = BOQItem(project_id=pid, **{k: v for k, v in data.items() if hasattr(BOQItem, k)})
    db.session.add(item); db.session.commit()
    return jsonify(item.to_dict()), 201

@boq_bp.route("/<int:pid>/bulk", methods=["POST"])
@jwt_required()
def bulk_add_boq(pid):
    err = admin_required()
    if err: return err
    items_data = request.get_json()
    created = []
    for i, d in enumerate(items_data):
        item = BOQItem(project_id=pid, sort_order=i,
                       **{k: v for k, v in d.items() if hasattr(BOQItem, k)})
        db.session.add(item); created.append(item)
    db.session.commit()
    return jsonify([i.to_dict() for i in created]), 201

@boq_bp.route("/<int:pid>/import-excel", methods=["POST"])
@jwt_required()
def import_boq_excel(pid):
    err = admin_required()
    if err: return err

    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "File must be .xlsx or .xls"}), 400

    try:
        import openpyxl
        # openpyxl only reads .xlsx — detect .xls and give a clear error
        filename = file.filename.lower()
        if filename.endswith('.xls') and not filename.endswith('.xlsx'):
            return jsonify({
                "error": "File is in old Excel 97-2003 (.xls) format. "
                         "Please re-save as Excel Workbook (.xlsx) in Excel: "
                         "File → Save As → Excel Workbook (.xlsx)"
            }), 400
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        err = str(e)
        if "xlrd" in err or "xls" in err.lower() or "OLE" in err:
            return jsonify({
                "error": "File appears to be in old .xls format. "
                         "Please re-save as Excel Workbook (.xlsx): "
                         "File → Save As → Excel Workbook (.xlsx)"
            }), 400
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    # Expected columns (matches BOQ_Import_Template / BEIL_WO249_EC_BOQ_Import):
    # A: Sr.No  B: Description  C: PO Qty  D: Unit  E: Rate  F: Amount
    # G: Site Zone  H: Item Type  I: Milestone Type  J: Remarks
    # Header is on row 3, hint row 4, data starts row 5
    HEADER_ROW = 3
    DATA_START_ROW = 5

    # Locate header row dynamically in case template varies slightly,
    # by scanning first 6 rows for a cell containing "Sr"
    header_row_idx = None
    for r in range(1, 7):
        val = ws.cell(row=r, column=1).value
        if val and "sr" in str(val).lower():
            header_row_idx = r
            break
    data_start = (header_row_idx + 2) if header_row_idx else DATA_START_ROW

    rows = []
    errors = []
    seen_sr = set()

    for r in range(data_start, ws.max_row + 1):
        sr_no = ws.cell(row=r, column=1).value
        description = ws.cell(row=r, column=2).value
        po_qty = ws.cell(row=r, column=3).value
        unit = ws.cell(row=r, column=4).value
        rate = ws.cell(row=r, column=5).value
        site_zone = ws.cell(row=r, column=7).value
        item_type = ws.cell(row=r, column=8).value

        # Stop at blank row or TOTAL row
        if sr_no is None and description is None:
            continue
        if sr_no and "total" in str(sr_no).lower():
            continue
        if not sr_no or not description:
            continue

        sr_no = str(sr_no).strip()
        description = str(description).strip()

        if sr_no in seen_sr:
            errors.append(f"Row {r}: duplicate Sr.No '{sr_no}' in file — skipped")
            continue
        seen_sr.add(sr_no)

        try:
            po_qty = float(po_qty) if po_qty not in (None, "") else 0
        except (ValueError, TypeError):
            errors.append(f"Row {r}: invalid PO Qty for '{sr_no}' — skipped")
            continue

        try:
            rate = float(rate) if rate not in (None, "") else 0
        except (ValueError, TypeError):
            rate = 0

        unit = str(unit).strip() if unit else "Nos."
        site_zone = str(site_zone).strip() if site_zone else "GENERAL"
        item_type = str(item_type).strip().lower() if item_type else "supply"
        # Accept both the internal value ("erection") and the display label
        # ("installation") users may type in the Excel — UI shows "Installation"
        # but the system stores "erection" internally everywhere.
        item_type_aliases = {
            "installation": "erection",
            "install": "erection",
            "erection": "erection",
            "supply": "supply",
            "commissioning": "commissioning",
            "commission": "commissioning",
        }
        item_type = item_type_aliases.get(item_type, "supply")

        rows.append({
            "sr_no": sr_no,
            "description": description,
            "po_qty": po_qty,
            "unit": unit,
            "rate": rate,
            "amount": round(po_qty * rate, 2),
            "site_zone": site_zone,
            "item_type": item_type,
        })

    if not rows:
        return jsonify({"error": "No valid rows found in file", "row_errors": errors}), 400

    # Check against existing Sr.Nos already in this project
    existing_sr = {i.sr_no for i in BOQItem.query.filter_by(project_id=pid, is_active=True).all()}
    duplicates_in_db = [r["sr_no"] for r in rows if r["sr_no"] in existing_sr]

    preview_only = request.args.get("preview", "false").lower() == "true"
    if preview_only:
        return jsonify({
            "rows": rows,
            "row_errors": errors,
            "duplicates_in_db": duplicates_in_db,
            "total_amount": sum(r["amount"] for r in rows),
        })

    # Actual import — skip any Sr.No that already exists in this project
    existing_count = BOQItem.query.filter_by(project_id=pid).count()
    created = []
    skipped = []
    for i, r in enumerate(rows):
        if r["sr_no"] in existing_sr:
            skipped.append(r["sr_no"])
            continue
        item = BOQItem(
            project_id=pid,
            sort_order=existing_count + i,
            sr_no=r["sr_no"], description=r["description"],
            po_qty=r["po_qty"], unit=r["unit"], rate=r["rate"],
            amount=r["amount"], site_zone=r["site_zone"], item_type=r["item_type"],
        )
        db.session.add(item)
        created.append(item)

    db.session.commit()

    return jsonify({
        "message": f"{len(created)} item(s) imported, {len(skipped)} skipped (already exist)",
        "imported_count": len(created),
        "skipped_sr_nos": skipped,
        "items": [i.to_dict() for i in created],
    }), 201


@boq_bp.route("/<int:pid>/export-excel", methods=["GET"])
@jwt_required()
def export_boq_excel(pid):
    """
    Export the project's current BOQ to Excel, in the SAME layout as the
    import template (title, hint row, headers on row 3, data from row 5).
    The exported file can be re-uploaded via Import Excel without any
    changes — enables a clean export -> edit -> re-import round trip.
    """
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = Project.query.get_or_404(pid)
    items = BOQItem.query.filter_by(project_id=pid, is_active=True)\
                          .order_by(BOQItem.sort_order, BOQItem.sr_no).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ Export"

    thin  = Side(style="thin")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    C     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    TEAL  = PatternFill("solid", fgColor="1D9E75")
    GRAY  = PatternFill("solid", fgColor="F1EFE8")
    GREEN = PatternFill("solid", fgColor="E1F5EE")
    AMBER = PatternFill("solid", fgColor="FFF8E1")

    def sc(r, c, val="", bold=False, fill=None, align=L, size=10, color="000000", fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, size=size, name="Arial",
                         color="FFFFFF" if fill == TEAL else color)
        if fill: cell.fill = fill
        cell.alignment = align
        cell.border = bdr
        if fmt: cell.number_format = fmt
        return cell

    col_widths = [12, 55, 10, 10, 14, 14, 18, 18, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1 — Title
    ws.merge_cells("A1:J1")
    sc(1, 1, f"BOQ EXPORT — {project.code} — {project.name}",
       bold=True, fill=TEAL, align=C, size=13)
    ws.row_dimensions[1].height = 28

    # Row 2 — Sub-header
    ws.merge_cells("A2:J2")
    sc(2, 1,
       "This file matches the Import template layout — edit and re-upload via "
       "Admin -> BOQ Manager -> Import Excel to update items in bulk.",
       fill=AMBER, align=L, size=9, color="854F0B")
    ws.row_dimensions[2].height = 18

    # Row 3 — Column headers (same order the import route reads)
    HEADERS = ["Sr. No. *", "Description *", "PO Qty *", "Unit *", "Rate (Rs.) *",
               "Amount (Rs.)", "Site Zone", "Item Type *", "Milestone Type", "Remarks"]
    for i, h in enumerate(HEADERS, 1):
        sc(3, i, h, bold=True, fill=TEAL, align=C, size=10)
    ws.row_dimensions[3].height = 28

    # Row 4 — Hints (kept identical to import template for consistency)
    HINTS = ["S-1, I-1A...", "Full item description", "Numeric (e.g. 4)",
              "Nos./Mtr/Set/Job", "Excl. GST", "Auto: Qty x Rate",
              "MPS SITE / STP SITE / SPS SITE / GENERAL",
              "supply / installation / commissioning", "standard", "Optional notes"]
    for i, h in enumerate(HINTS, 1):
        sc(4, i, h, fill=GRAY, align=L, size=8, color="5F5E5A")
    ws.row_dimensions[4].height = 22

    # Row 5+ — actual current BOQ data
    r = 5
    for item in items:
        vals = [
            item.sr_no, item.description, float(item.po_qty), item.unit,
            float(item.rate), float(item.amount), item.site_zone or "GENERAL",
            item.item_type, item.milestone_type or "standard", "",
        ]
        for i, v in enumerate(vals, 1):
            fmt = "#,##0.00" if i in (3, 5, 6) else None
            align = C if i not in (2,) else L
            sc(r, i, v, fill=GREEN, align=align, fmt=fmt)
        ws.row_dimensions[r].height = 20
        r += 1

    # Total row
    ws.merge_cells(f"A{r}:E{r}")
    sc(r, 1, f"TOTAL — {len(items)} item(s)", bold=True, fill=GRAY, align=C)
    ws.cell(row=r, column=6).value = sum(float(i.amount) for i in items)
    ws.cell(row=r, column=6).number_format = "#,##0.00"
    ws.cell(row=r, column=6).font = Font(bold=True, size=10, name="Arial")
    ws.cell(row=r, column=6).fill = GREEN
    ws.cell(row=r, column=6).alignment = C
    ws.cell(row=r, column=6).border = bdr
    for c in range(7, 11):
        sc(r, c, "", fill=GRAY)

    ws.freeze_panes = "A5"

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True,
                     download_name=f"BOQ_Export_{project.code}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@boq_bp.route("/item/<int:iid>", methods=["PUT"])
@jwt_required()
def update_boq_item(iid):
    err = admin_required()
    if err: return err
    item = BOQItem.query.get_or_404(iid)
    data = request.get_json()
    for k, v in data.items():
        if hasattr(item, k): setattr(item, k, v)
    db.session.commit()
    return jsonify(item.to_dict())

@boq_bp.route("/item/<int:iid>", methods=["DELETE"])
@jwt_required()
def delete_boq_item(iid):
    err = admin_required()
    if err: return err
    item = BOQItem.query.get_or_404(iid)
    item.is_active = False
    db.session.commit()
    return jsonify({"message": "Deleted"})

# ── Users ─────────────────────────────────────────────────────────────────────

users_bp = Blueprint("users", __name__)

@users_bp.route("/", methods=["GET"])
@jwt_required()
def list_users():
    err = admin_required()
    if err: return err
    users = User.query.order_by(User.name).all()
    return jsonify([u.to_dict() for u in users])

@users_bp.route("/", methods=["POST"])
@jwt_required()
def create_user():
    err = admin_required()
    if err: return err
    data = request.get_json()
    if User.query.filter_by(email=data["email"].lower()).first():
        return jsonify({"error": "Email already exists"}), 400
    u = User(name=data["name"], email=data["email"].lower(),
             role=data["role"], phone_whatsapp=data.get("phone_whatsapp",""))
    u.set_password(data["password"])
    db.session.add(u); db.session.commit()
    return jsonify(u.to_dict()), 201

@users_bp.route("/<int:uid>", methods=["PUT"])
@jwt_required()
def update_user(uid):
    me = current_user()
    if me.role != "admin" and me.id != uid:
        return jsonify({"error": "Forbidden"}), 403
    u = User.query.get_or_404(uid)
    data = request.get_json()
    allowed = ["name","phone_whatsapp","notify_grn","notify_dispatch","notify_progress","notify_ra"]
    if me.role == "admin": allowed += ["role","is_active","email"]
    for k in allowed:
        if k in data: setattr(u, k, data[k])
    if me.role == "admin" and data.get("password"):
        u.set_password(data["password"])
    db.session.commit()
    return jsonify(u.to_dict())

@users_bp.route("/whatsapp-contacts", methods=["GET"])
@jwt_required()
def wa_contacts():
    users = User.query.filter(User.is_active==True, User.phone_whatsapp!="").all()
    return jsonify([{"id": u.id, "name": u.name, "role": u.role,
                     "phone": u.phone_whatsapp} for u in users])

# ── GRN ──────────────────────────────────────────────────────────────────────

grn_bp = Blueprint("grn", __name__)

def next_grn_number(pid):
    count = GRN.query.filter_by(project_id=pid).count()
    return f"GRN-{pid:03d}-{count+1:04d}"

@grn_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def list_grns(pid):
    grns = GRN.query.filter_by(project_id=pid).order_by(GRN.created_at.desc()).all()
    return jsonify([g.to_dict() for g in grns])

@grn_bp.route("/<int:pid>", methods=["POST"])
@jwt_required()
def create_grn(pid):
    data = request.get_json()
    project = Project.query.get_or_404(pid)
    grn = GRN(
        project_id=pid,
        grn_number=next_grn_number(pid),
        grn_date=date.fromisoformat(data["grn_date"]),
        boq_item_id=data["boq_item_id"],
        qty_received=data["qty_received"],
        vendor_name=data.get("vendor_name",""),
        challan_no=data.get("challan_no",""),
        vehicle_no=data.get("vehicle_no",""),
        remarks=data.get("remarks",""),
        created_by=int(get_jwt_identity())
    )
    db.session.add(grn); db.session.commit()
    try: notify_grn_created(grn, project)
    except Exception as e: print(f"Notify error: {e}")
    return jsonify(grn.to_dict()), 201

@grn_bp.route("/<int:gid>", methods=["GET"])
@jwt_required()
def get_grn(gid):
    return jsonify(GRN.query.get_or_404(gid).to_dict())

# ── Dispatch ──────────────────────────────────────────────────────────────────

dispatch_bp = Blueprint("dispatch", __name__)

def next_dn_number(pid):
    count = DispatchNote.query.filter_by(project_id=pid).count()
    return f"DN-{pid:03d}-{count+1:04d}"

@dispatch_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def list_dispatches(pid):
    dns = DispatchNote.query.filter_by(project_id=pid).order_by(DispatchNote.created_at.desc()).all()
    return jsonify([d.to_dict() for d in dns])

@dispatch_bp.route("/<int:pid>", methods=["POST"])
@jwt_required()
def create_dispatch(pid):
    data = request.get_json()
    project = Project.query.get_or_404(pid)
    dn = DispatchNote(
        project_id=pid,
        dn_number=next_dn_number(pid),
        dispatch_date=date.fromisoformat(data["dispatch_date"]),
        boq_item_id=data["boq_item_id"],
        qty_dispatched=data["qty_dispatched"],
        site_destination=data.get("site_destination",""),
        vehicle_no=data.get("vehicle_no",""),
        driver_name=data.get("driver_name",""),
        lr_number=data.get("lr_number",""),
        remarks=data.get("remarks",""),
        created_by=int(get_jwt_identity())
    )
    db.session.add(dn); db.session.commit()
    try: notify_dispatch_created(dn, project)
    except Exception as e: print(f"Notify error: {e}")
    return jsonify(dn.to_dict()), 201

@dispatch_bp.route("/pending-invoice/<int:pid>", methods=["GET"])
@jwt_required()
def pending_invoice(pid):
    dns = DispatchNote.query.filter_by(project_id=pid, invoice_status="pending").all()
    return jsonify([d.to_dict() for d in dns])

@dispatch_bp.route("/<int:did>/mark-invoiced", methods=["PUT"])
@jwt_required()
def mark_invoiced(did):
    dn = DispatchNote.query.get_or_404(did)
    dn.invoice_status = "invoiced"
    db.session.commit()
    return jsonify(dn.to_dict())

# ── Site Progress ─────────────────────────────────────────────────────────────

site_bp = Blueprint("site", __name__)

@site_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def list_progress(pid):
    boq = BOQItem.query.filter_by(project_id=pid, is_active=True)\
                        .order_by(BOQItem.sort_order).all()
    result = []
    for item in boq:
        entries = SiteProgress.query.filter_by(project_id=pid, boq_item_id=item.id)\
                                    .order_by(SiteProgress.progress_date.desc()).all()
        total_inst = sum(float(e.qty_installed) for e in entries)
        total_comm = sum(float(e.qty_commissioned) for e in entries)
        result.append({
            **item.to_dict(),
            "total_installed": total_inst,
            "total_commissioned": total_comm,
            "pct_installed": round(total_inst / float(item.po_qty) * 100, 1) if float(item.po_qty) else 0,
            "pct_commissioned": round(total_comm / float(item.po_qty) * 100, 1) if float(item.po_qty) else 0,
            "history": [e.to_dict() for e in entries[:5]],
        })
    return jsonify(result)

@site_bp.route("/<int:pid>/update", methods=["POST"])
@jwt_required()
def update_progress(pid):
    data = request.get_json()
    project = Project.query.get_or_404(pid)
    user = current_user()
    updates = data.get("updates", [])
    count = 0
    for upd in updates:
        if upd.get("qty_installed", 0) == 0 and upd.get("qty_commissioned", 0) == 0:
            continue
        sp = SiteProgress(
            project_id=pid,
            boq_item_id=upd["boq_item_id"],
            progress_date=date.fromisoformat(upd.get("progress_date", date.today().isoformat())),
            qty_installed=upd.get("qty_installed", 0),
            qty_commissioned=upd.get("qty_commissioned", 0),
            notes=upd.get("notes", ""),
            updated_by=user.id
        )
        db.session.add(sp); count += 1
    db.session.commit()
    if count > 0:
        try: notify_progress_updated(project, user.name, count)
        except Exception as e: print(f"Notify error: {e}")
    return jsonify({"message": f"{count} items updated"})

# ── RA Bill ───────────────────────────────────────────────────────────────────

ra_bp = Blueprint("ra", __name__)

@ra_bp.route("/<int:pid>", methods=["GET"])
@jwt_required()
def list_ra(pid):
    bills = RABill.query.filter_by(project_id=pid).order_by(RABill.ra_number.desc()).all()
    return jsonify([b.to_dict() for b in bills])

@ra_bp.route("/<int:pid>/compute", methods=["POST"])
@jwt_required()
def compute_ra(pid):
    """Compute RA bill values from site progress — returns preview, does not save."""
    project = Project.query.get_or_404(pid)
    data = request.get_json()
    invoice_date = date.fromisoformat(data.get("invoice_date", date.today().isoformat()))

    boq_items = BOQItem.query.filter_by(project_id=pid, is_active=True).all()
    prev_ra = RABill.query.filter_by(project_id=pid)\
                          .order_by(RABill.ra_number.desc()).first()

    lines = []
    supply_this = Decimal("0"); ec_this = Decimal("0")

    for item in boq_items:
        all_progress = SiteProgress.query.filter_by(project_id=pid, boq_item_id=item.id).all()
        total_installed   = sum(Decimal(str(e.qty_installed)) for e in all_progress)
        total_commissioned = sum(Decimal(str(e.qty_commissioned)) for e in all_progress)

        if item.item_type in ["supply", "erection"]:
            qty_upto = total_installed
        else:
            qty_upto = total_commissioned

        # Get qty from previous RA
        if prev_ra:
            prev_line = RABillLine.query.filter_by(ra_bill_id=prev_ra.id, boq_item_id=item.id).first()
            qty_prev = Decimal(str(prev_line.qty_upto)) if prev_line else Decimal("0")
        else:
            qty_prev = Decimal("0")

        qty_this    = max(qty_upto - qty_prev, Decimal("0"))
        qty_balance = max(Decimal(str(item.po_qty)) - qty_upto, Decimal("0"))
        rate        = Decimal(str(item.rate))

        amount_prev    = qty_prev    * rate
        amount_this    = qty_this    * rate
        amount_upto    = qty_upto    * rate
        amount_balance = qty_balance * rate

        if item.item_type == "supply":
            supply_this += amount_this
        else:
            ec_this += amount_this

        lines.append({
            "boq_item_id": item.id, "sr_no": item.sr_no,
            "description": item.description, "unit": item.unit,
            "rate": float(rate), "po_qty": float(item.po_qty),
            "qty_prev": float(qty_prev), "qty_this": float(qty_this),
            "qty_upto": float(qty_upto), "qty_balance": float(qty_balance),
            "amount_prev": float(amount_prev), "amount_this": float(amount_this),
            "amount_upto": float(amount_upto), "amount_balance": float(amount_balance),
            "item_type": item.item_type,
        })

    taxable    = supply_this + ec_this
    igst       = (taxable * Decimal(str(project.igst_rate)) / 100).quantize(Decimal("0.01"))
    cgst       = (taxable * Decimal(str(project.cgst_rate)) / 100).quantize(Decimal("0.01"))
    sgst       = (taxable * Decimal(str(project.sgst_rate)) / 100).quantize(Decimal("0.01"))
    gross      = taxable + igst + cgst + sgst
    adv_rec    = (supply_this * Decimal(str(project.pt_advance_pct)) / 100).quantize(Decimal("0.01"))
    retention  = (taxable * Decimal(str(project.pt_retention_pct)) / 100).quantize(Decimal("0.01"))
    net        = gross - adv_rec - retention

    ra_no = (prev_ra.ra_number + 1) if prev_ra else 1

    return jsonify({
        "ra_number": ra_no,
        "invoice_no": f"{project.invoice_prefix}/{ra_no:03d}",
        "invoice_date": invoice_date.isoformat(),
        "supply_value_this": float(supply_this),
        "ec_value_this": float(ec_this),
        "taxable_value": float(taxable),
        "igst_amount": float(igst),
        "cgst_amount": float(cgst),
        "sgst_amount": float(sgst),
        "gross_total": float(gross),
        "advance_recovery": float(adv_rec),
        "retention_deduction": float(retention),
        "net_payable": float(net),
        "lines": lines,
    })

@ra_bp.route("/<int:pid>/save", methods=["POST"])
@jwt_required()
def save_ra(pid):
    project = Project.query.get_or_404(pid)
    data = request.get_json()
    user = current_user()

    ra = RABill(
        project_id=pid,
        ra_number=data["ra_number"],
        invoice_no=data["invoice_no"],
        invoice_date=date.fromisoformat(data["invoice_date"]),
        supply_value_this=data["supply_value_this"],
        ec_value_this=data["ec_value_this"],
        taxable_value=data["taxable_value"],
        igst_amount=data["igst_amount"],
        cgst_amount=data["cgst_amount"],
        sgst_amount=data["sgst_amount"],
        gross_total=data["gross_total"],
        advance_recovery=data["advance_recovery"],
        retention_deduction=data.get("retention_deduction", 0),
        net_payable=data["net_payable"],
        notes=data.get("notes",""),
        created_by=user.id,
        status="draft",
    )
    db.session.add(ra); db.session.flush()

    for li in data.get("lines", []):
        line = RABillLine(
            ra_bill_id=ra.id,
            boq_item_id=li["boq_item_id"],
            qty_prev=li["qty_prev"], qty_this=li["qty_this"],
            qty_upto=li["qty_upto"], qty_balance=li["qty_balance"],
            amount_prev=li["amount_prev"], amount_this=li["amount_this"],
            amount_upto=li["amount_upto"], amount_balance=li["amount_balance"],
        )
        db.session.add(line)

    project.current_ra_no = data["ra_number"] + 1
    db.session.commit()
    return jsonify(ra.to_dict()), 201

@ra_bp.route("/<int:rid>/export/excel", methods=["GET"])
@jwt_required()
def export_excel(rid):
    from flask import send_file
    ra = RABill.query.get_or_404(rid)
    project = Project.query.get(ra.project_id)
    lines = [l.to_dict() for l in ra.line_items]
    path = generate_ra_excel(ra, project, lines)
    return send_file(path, as_attachment=True,
                     download_name=f"RA_{ra.ra_number}_{ra.invoice_no}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@ra_bp.route("/<int:rid>/export/pdf", methods=["GET"])
@jwt_required()
def export_pdf(rid):
    from flask import send_file
    ra = RABill.query.get_or_404(rid)
    project = Project.query.get(ra.project_id)
    lines = [l.to_dict() for l in ra.line_items]
    path = generate_ra_pdf(ra, project, lines)
    if not path:
        return jsonify({"error": "PDF generation failed — WeasyPrint not available"}), 500
    return send_file(path, as_attachment=True,
                     download_name=f"RA_{ra.ra_number}_{ra.invoice_no}.pdf",
                     mimetype="application/pdf")

@ra_bp.route("/<int:rid>/status", methods=["PUT"])
@jwt_required()
def update_status(rid):
    ra = RABill.query.get_or_404(rid)
    data = request.get_json()
    ra.status = data.get("status", ra.status)
    if data.get("irn_number"): ra.irn_number = data["irn_number"]
    if ra.status == "submitted": ra.submitted_at = datetime.utcnow()
    db.session.commit()
    if ra.status == "submitted":
        project = Project.query.get(ra.project_id)
        try: notify_ra_generated(ra, project)
        except Exception as e: print(f"Notify error: {e}")
    return jsonify(ra.to_dict())

# ── Notifications ─────────────────────────────────────────────────────────────

notif_bp = Blueprint("notifications", __name__)

@notif_bp.route("/", methods=["GET"])
@jwt_required()
def list_notifs():
    uid = int(get_jwt_identity())
    pid = request.args.get("project_id")
    q = Notification.query.filter_by(user_id=uid)
    if pid: q = q.filter_by(project_id=int(pid))
    notifs = q.order_by(Notification.created_at.desc()).limit(50).all()
    return jsonify([n.to_dict() for n in notifs])

@notif_bp.route("/<int:nid>/read", methods=["PUT"])
@jwt_required()
def mark_read(nid):
    n = Notification.query.get_or_404(nid)
    n.is_read = True; db.session.commit()
    return jsonify(n.to_dict())

@notif_bp.route("/read-all", methods=["PUT"])
@jwt_required()
def mark_all_read():
    uid = int(get_jwt_identity())
    Notification.query.filter_by(user_id=uid, is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"message": "All marked read"})
